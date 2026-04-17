from __future__ import annotations

from datetime import UTC, datetime
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

from englishbot.application.media_catalog_use_cases import (
    TOPICS_HEADERS,
    TOPICS_SHEET,
    WORDS_IN_TOPICS_HEADERS,
    WORDS_IN_TOPICS_SHEET,
    ExportMediaCatalogWorkbookUseCase,
    ImportMediaCatalogWorkbookUseCase,
    SaveCatalogUploadedImageUseCase,
)
from englishbot.infrastructure.sqlite_store import SQLiteContentStore


class FakeRemoteImageDownloader:
    def __init__(self) -> None:
        self.calls: list[tuple[str, Path]] = []

    def download(self, *, url: str, output_path: Path) -> None:
        self.calls.append((url, output_path))
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"fake-image")


def _build_store(tmp_path: Path) -> SQLiteContentStore:
    store = SQLiteContentStore(db_path=tmp_path / "data" / "englishbot.db")
    store.upsert_content_pack(
        {
            "topic": {"id": "fairy-tales", "title": "Fairy Tales"},
            "lessons": [{"id": "fairy-tales-1", "title": "Lesson 1"}],
            "vocabulary_items": [
                {
                    "id": "fairy-tales-fairy",
                    "english_word": "fairy",
                    "translation": "фея",
                    "image_ref": "https://img.example/fairy.png",
                    "image_source": "https://img.example/fairy-preview.png",
                    "image_prompt": "fairy in forest",
                    "pixabay_search_query": "fairy illustration",
                    "source_fragment": "fairy tale fragment",
                }
            ],
        }
    )
    store.upsert_content_pack(
        {
            "topic": {"id": "cleaning-stuff", "title": "Cleaning Stuff"},
            "lessons": [],
            "vocabulary_items": [
                {
                    "id": "cleaning-stuff-broom",
                    "english_word": "broom",
                    "translation": "метла",
                }
            ],
        }
    )
    return store


def test_export_media_catalog_workbook_writes_two_simple_sheets(tmp_path: Path) -> None:
    store = _build_store(tmp_path)
    output_path = tmp_path / "exports" / "catalog.xlsx"

    local_asset_path = tmp_path / "assets" / "cleaning-stuff" / "cleaning-stuff-broom.png"
    local_asset_path.parent.mkdir(parents=True, exist_ok=True)
    Image.new("RGB", (640, 480), color=(50, 100, 150)).save(local_asset_path)
    workbook_source_pack = store.get_content_pack("cleaning-stuff")
    workbook_source_pack["vocabulary_items"][0]["image_ref"] = local_asset_path.as_posix()
    store.upsert_content_pack(workbook_source_pack)

    result = ExportMediaCatalogWorkbookUseCase(
        store=store,
        assets_dir=tmp_path / "assets",
        web_app_base_url="https://admin.example.com",
        public_asset_signing_secret="preview-secret",
    ).execute(output_path=output_path)

    assert output_path.is_file()
    assert len(result.topics) == 2
    workbook = load_workbook(output_path)
    assert workbook.sheetnames == [TOPICS_SHEET, WORDS_IN_TOPICS_SHEET]
    topics_sheet = workbook[TOPICS_SHEET]
    words_sheet = workbook[WORDS_IN_TOPICS_SHEET]
    assert tuple(cell.value for cell in topics_sheet[1][: len(TOPICS_HEADERS)]) == TOPICS_HEADERS
    assert tuple(cell.value for cell in words_sheet[1][: len(WORDS_IN_TOPICS_HEADERS)]) == WORDS_IN_TOPICS_HEADERS
    assert topics_sheet["A2"].value == "Cleaning Stuff"
    assert words_sheet["A2"].value == "Cleaning Stuff"
    assert words_sheet["B2"].value == "broom"
    assert words_sheet["E2"].value == '=IF(F2="","",IMAGE(F2))'
    assert str(words_sheet["F2"].value).startswith("https://admin.example.com/public-assets/file?path=")
    validations = list(words_sheet.data_validations.dataValidation)
    assert len(validations) == 1
    assert validations[0].formula1 == "=topics!$A$2:$A$3"
    assert "A2:A200" in str(validations[0].sqref)


def test_import_media_catalog_workbook_uses_unique_topic_titles_and_word_matching(tmp_path: Path) -> None:
    store = _build_store(tmp_path)
    downloader = FakeRemoteImageDownloader()
    workbook_path = tmp_path / "exports" / "catalog.xlsx"
    ExportMediaCatalogWorkbookUseCase(store=store).execute(output_path=workbook_path)

    workbook = load_workbook(workbook_path)
    topics_sheet = workbook[TOPICS_SHEET]
    words_sheet = workbook[WORDS_IN_TOPICS_SHEET]
    topics_sheet.append(["Products"])
    words_sheet["C2"] = "метёлка"
    words_sheet.append(
        [
            "Cleaning Stuff",
            "fairy",
            "Fairy",
            "",
            "",
            "https://img.example/detergent-fairy.png",
            "https://img.example/detergent-fairy-preview.png",
            "detergent bottle",
            "fairy detergent",
            "cleaning fragment",
            "TRUE",
        ]
    )
    words_sheet.append(
        [
            "Products",
            "soap",
            "мыло",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "TRUE",
        ]
    )
    workbook.save(workbook_path)

    result = ImportMediaCatalogWorkbookUseCase(
        store=store,
        assets_dir=tmp_path / "assets",
        remote_image_downloader=downloader,
    ).execute(input_path=workbook_path)

    assert result.updated_count == 4
    assert result.topic_count == 3
    assert result.backup_path is not None
    assert result.backup_path.is_file()
    cleaning_items = store.list_vocabulary_by_topic("cleaning-stuff")
    broom_item = next(item for item in cleaning_items if item.english_word == "broom")
    fairy_item = next(item for item in cleaning_items if item.english_word == "fairy")
    assert broom_item.translation == "метёлка"
    assert fairy_item.image_ref == (tmp_path / "assets" / "cleaning-stuff" / "cleaning-stuff-fairy.png").as_posix()
    assert fairy_item.image_source == "https://img.example/detergent-fairy-preview.png"
    assert downloader.calls == [
        (
            "https://img.example/detergent-fairy.png",
            tmp_path / "assets" / "cleaning-stuff" / "cleaning-stuff-fairy.png",
        ),
        (
            "https://img.example/fairy.png",
            tmp_path / "assets" / "fairy-tales" / "fairy-tales-fairy.png",
        ),
    ] or downloader.calls == [
        (
            "https://img.example/fairy.png",
            tmp_path / "assets" / "fairy-tales" / "fairy-tales-fairy.png",
        ),
        (
            "https://img.example/detergent-fairy.png",
            tmp_path / "assets" / "cleaning-stuff" / "cleaning-stuff-fairy.png",
        ),
    ]
    assert store.get_topic("products") is not None
    assert store.list_vocabulary_by_topic("products")[0].english_word == "soap"
    lexeme = store.get_lexeme_by_normalized_headword("fairy")
    assert lexeme is not None
    assert store.get_content_pack("fairy-tales")["lessons"] == [{"id": "fairy-tales-1", "title": "Lesson 1"}]


def test_import_media_catalog_workbook_preserves_signed_local_asset_reference_without_redownload(
    tmp_path: Path,
) -> None:
    store = _build_store(tmp_path)
    downloader = FakeRemoteImageDownloader()
    local_asset_path = tmp_path / "assets" / "cleaning-stuff" / "cleaning-stuff-broom.png"
    local_asset_path.parent.mkdir(parents=True, exist_ok=True)
    Image.new("RGB", (320, 240), color=(10, 20, 30)).save(local_asset_path)
    pack = store.get_content_pack("cleaning-stuff")
    pack["vocabulary_items"][0]["image_ref"] = local_asset_path.as_posix()
    store.upsert_content_pack(pack)
    workbook_path = tmp_path / "exports" / "catalog.xlsx"

    ExportMediaCatalogWorkbookUseCase(
        store=store,
        assets_dir=tmp_path / "assets",
        web_app_base_url="https://admin.example.com",
        public_asset_signing_secret="preview-secret",
    ).execute(output_path=workbook_path)

    result = ImportMediaCatalogWorkbookUseCase(
        store=store,
        assets_dir=tmp_path / "assets",
        web_app_base_url="https://admin.example.com",
        public_asset_signing_secret="preview-secret",
        remote_image_downloader=downloader,
    ).execute(input_path=workbook_path)

    assert result.updated_count == 2
    assert downloader.calls == [
        (
            "https://img.example/fairy.png",
            tmp_path / "assets" / "fairy-tales" / "fairy-tales-fairy.png",
        )
    ]
    broom_item = next(item for item in store.list_vocabulary_by_topic("cleaning-stuff") if item.english_word == "broom")
    assert broom_item.image_ref == local_asset_path.as_posix()


def test_import_media_catalog_workbook_rolls_back_all_topics_on_apply_failure(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    store = _build_store(tmp_path)
    downloader = FakeRemoteImageDownloader()
    workbook_path = tmp_path / "exports" / "catalog.xlsx"
    ExportMediaCatalogWorkbookUseCase(store=store).execute(output_path=workbook_path)

    workbook = load_workbook(workbook_path)
    topics_sheet = workbook[TOPICS_SHEET]
    words_sheet = workbook[WORDS_IN_TOPICS_SHEET]
    topics_sheet.append(["Products"])
    words_sheet["C2"] = "метёлка"
    words_sheet.append(
        [
            "Products",
            "soap",
            "мыло",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "TRUE",
        ]
    )
    workbook.save(workbook_path)

    original_helper = store._upsert_content_pack_with_connection
    call_count = 0

    def failing_helper(connection, content_pack):
        nonlocal call_count
        call_count += 1
        result = original_helper(connection, content_pack)
        if call_count == 2:
            raise RuntimeError("boom during second topic apply")
        return result

    monkeypatch.setattr(store, "_upsert_content_pack_with_connection", failing_helper)

    with pytest.raises(RuntimeError, match="boom during second topic apply"):
        ImportMediaCatalogWorkbookUseCase(
            store=store,
            assets_dir=tmp_path / "assets",
            remote_image_downloader=downloader,
        ).execute(input_path=workbook_path)

    cleaning_items = store.list_vocabulary_by_topic("cleaning-stuff")
    broom_item = next(item for item in cleaning_items if item.english_word == "broom")
    assert broom_item.translation == "метла"
    assert store.get_topic("products") is None


def test_save_catalog_uploaded_image_use_case_stores_file_and_returns_public_url(tmp_path: Path) -> None:
    input_path = tmp_path / "incoming.png"
    input_path.write_bytes(b"image-bytes")

    result = SaveCatalogUploadedImageUseCase(
        assets_dir=tmp_path / "assets",
        web_app_base_url="https://admin.example.com",
        public_asset_signing_secret="preview-secret",
    ).execute(
        input_path=input_path,
        original_file_name="dragon.png",
        mime_type="image/png",
    )

    assert result.asset_path.is_file()
    assert result.asset_path.read_bytes() == b"image-bytes"
    assert result.asset_path.parent.name == datetime.now(UTC).strftime("%Y%m%d")
    assert str(result.public_url).startswith("https://admin.example.com/public-assets/file?path=catalog-uploads/")
