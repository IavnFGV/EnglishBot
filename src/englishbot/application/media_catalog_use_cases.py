from __future__ import annotations

import re
import sqlite3
from datetime import UTC, datetime
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font

from englishbot.domain.models import Topic
from englishbot.image_generation.pixabay import RemoteImageDownloader
from englishbot.infrastructure.sqlite_store import SQLiteContentStore
from englishbot.logging_utils import logged_service_call
from englishbot.public_assets import build_public_asset_file_url, resolve_signed_public_asset_file_path

TOPICS_SHEET = "topics"
WORDS_IN_TOPICS_SHEET = "words_in_topics"

TOPICS_HEADERS = ("topic_title",)
WORDS_IN_TOPICS_HEADERS = (
    "topic_title",
    "english_word",
    "translation",
    "meaning_hint",
    "preview",
    "image_ref",
    "image_source",
    "image_prompt",
    "pixabay_search_query",
    "source_fragment",
    "is_active",
)


@dataclass(frozen=True, slots=True)
class WordInTopicRow:
    topic_title: str
    english_word: str
    translation: str
    meaning_hint: str | None
    image_ref: str | None
    image_source: str | None
    image_prompt: str | None
    pixabay_search_query: str | None
    source_fragment: str | None
    is_active: bool


@dataclass(frozen=True, slots=True)
class CatalogWorkbookExport:
    topics: tuple[Topic, ...]
    words_in_topics: tuple[WordInTopicRow, ...]


@dataclass(frozen=True, slots=True)
class MediaCatalogImportResult:
    updated_count: int
    topic_count: int
    backup_path: Path | None = None


class ExportMediaCatalogWorkbookUseCase:
    def __init__(
        self,
        *,
        store: SQLiteContentStore,
        assets_dir: Path = Path("assets"),
        web_app_base_url: str = "",
        public_asset_signing_secret: str = "",
    ) -> None:
        self._store = store
        self._assets_dir = assets_dir
        self._web_app_base_url = web_app_base_url
        self._public_asset_signing_secret = public_asset_signing_secret

    @logged_service_call(
        "ExportMediaCatalogWorkbookUseCase.execute",
        include=("topic_id",),
        result=lambda value: {
            "topic_count": len(value.topics),
            "word_row_count": len(value.words_in_topics),
        },
    )
    def execute(self, *, output_path: Path, topic_id: str | None = None) -> CatalogWorkbookExport:
        export = self._build_export(topic_id=topic_id)
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        self._write_sheet(
            workbook=workbook,
            sheet_name=TOPICS_SHEET,
            headers=TOPICS_HEADERS,
            rows=[(topic.title,) for topic in export.topics],
        )
        self._write_sheet(
            workbook=workbook,
            sheet_name=WORDS_IN_TOPICS_SHEET,
            headers=WORDS_IN_TOPICS_HEADERS,
            rows=[self._worksheet_row_for_item(row=row, row_index=index + 2) for index, row in enumerate(export.words_in_topics)],
        )
        self._add_topic_title_validation(
            workbook=workbook,
            topic_count=len(export.topics),
            word_row_count=max(len(export.words_in_topics), 1),
        )
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return export

    def _worksheet_row_for_item(self, *, row: WordInTopicRow, row_index: int) -> tuple[object, ...]:
        exported_image_ref = _exported_image_ref_for_item(
            image_ref=row.image_ref,
            assets_dir=self._assets_dir,
            web_app_base_url=self._web_app_base_url,
            public_asset_signing_secret=self._public_asset_signing_secret,
        )
        return (
            row.topic_title,
            row.english_word,
            row.translation,
            row.meaning_hint,
            _preview_formula_for_cell(row_index=row_index, image_ref_column_index=6)
            if _is_remote_url(exported_image_ref or "")
            else None,
            exported_image_ref,
            row.image_source,
            row.image_prompt,
            row.pixabay_search_query,
            row.source_fragment,
            _bool_cell_value(row.is_active),
        )

    def _build_export(self, *, topic_id: str | None) -> CatalogWorkbookExport:
        topics = tuple(
            topic
            for topic in self._store.list_topics()
            if topic_id is None or topic.id == topic_id
        )
        topic_title_map = {topic.id: topic.title for topic in topics}
        items = (
            self._store.list_vocabulary_by_topic(topic_id)
            if topic_id is not None
            else self._store.list_all_vocabulary()
        )
        rows = tuple(
            WordInTopicRow(
                topic_title=topic_title_map[resolved_topic_id],
                english_word=item.english_word,
                translation=item.translation,
                meaning_hint=item.meaning_hint,
                image_ref=item.image_ref,
                image_source=item.image_source,
                image_prompt=item.image_prompt,
                pixabay_search_query=item.pixabay_search_query,
                source_fragment=item.source_fragment,
                is_active=item.is_active,
            )
            for item in items
            for resolved_topic_id in [_resolve_topic_id(self._store, item.id)]
            if resolved_topic_id in topic_title_map
        )
        return CatalogWorkbookExport(topics=topics, words_in_topics=rows)

    def _write_sheet(
        self,
        *,
        workbook: Workbook,
        sheet_name: str,
        headers: tuple[str, ...],
        rows: list[tuple[object, ...]],
    ) -> None:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=index, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center")
        for column, width in _column_widths_for_headers(headers).items():
            sheet.column_dimensions[column].width = width
        for row_index, row_values in enumerate(rows, start=2):
            for column_index, value in enumerate(row_values, start=1):
                sheet.cell(row=row_index, column=column_index, value=value)
            if sheet_name == WORDS_IN_TOPICS_SHEET and row_values[4]:
                sheet.row_dimensions[row_index].height = 96

    def _add_topic_title_validation(
        self,
        *,
        workbook: Workbook,
        topic_count: int,
        word_row_count: int,
    ) -> None:
        if topic_count <= 0:
            return
        words_sheet = workbook[WORDS_IN_TOPICS_SHEET]
        max_row = max(word_row_count + 25, 200)
        validation = DataValidation(
            type="list",
            formula1=f"={TOPICS_SHEET}!$A$2:$A${topic_count + 1}",
            allow_blank=False,
        )
        validation.promptTitle = "Choose topic"
        validation.prompt = "Choose a topic title from the topics sheet."
        validation.errorTitle = "Invalid topic"
        validation.error = "Select a topic title from the dropdown list."
        words_sheet.add_data_validation(validation)
        validation.add(f"A2:A{max_row}")


class ImportMediaCatalogWorkbookUseCase:
    def __init__(
        self,
        *,
        store: SQLiteContentStore,
        assets_dir: Path = Path("assets"),
        web_app_base_url: str = "",
        public_asset_signing_secret: str = "",
        remote_image_downloader: RemoteImageDownloader | None = None,
    ) -> None:
        self._store = store
        self._assets_dir = assets_dir
        self._web_app_base_url = web_app_base_url
        self._public_asset_signing_secret = public_asset_signing_secret
        self._remote_image_downloader = remote_image_downloader or RemoteImageDownloader()

    @logged_service_call(
        "ImportMediaCatalogWorkbookUseCase.execute",
        include=("topic_id",),
        result=lambda value: {
            "updated_count": value.updated_count,
            "topic_count": value.topic_count,
            "backup_created": value.backup_path is not None,
        },
    )
    def execute(self, *, input_path: Path, topic_id: str | None = None) -> MediaCatalogImportResult:
        workbook = load_workbook(input_path)
        if TOPICS_SHEET not in workbook.sheetnames:
            raise ValueError(f"Workbook must contain '{TOPICS_SHEET}' sheet.")
        if WORDS_IN_TOPICS_SHEET not in workbook.sheetnames:
            raise ValueError(f"Workbook must contain '{WORDS_IN_TOPICS_SHEET}' sheet.")

        topics_sheet = workbook[TOPICS_SHEET]
        words_sheet = workbook[WORDS_IN_TOPICS_SHEET]
        _validate_headers(topics_sheet, TOPICS_HEADERS, sheet_name=TOPICS_SHEET)
        _validate_headers(words_sheet, WORDS_IN_TOPICS_HEADERS, sheet_name=WORDS_IN_TOPICS_SHEET)

        workbook_topics = _read_workbook_topics(topics_sheet)
        existing_topics_by_title = {topic.title.casefold(): topic for topic in self._store.list_topics()}

        rows_by_topic_title: dict[str, list[dict[str, object]]] = defaultdict(list)
        for row in words_sheet.iter_rows(min_row=2, values_only=True):
            topic_title = _normalized_optional_str(row[0])
            english_word = _normalized_optional_str(row[1])
            translation = _normalized_optional_str(row[2])
            if topic_title is None and english_word is None and translation is None:
                continue
            if topic_title is None or topic_title.casefold() not in workbook_topics:
                raise ValueError("Each words_in_topics row must reference a topic_title from the topics sheet.")
            if english_word is None or translation is None:
                raise ValueError(
                    f"Row for topic '{topic_title}' must include both english_word and translation."
                )
            matched_topic = existing_topics_by_title.get(topic_title.casefold())
            if topic_id is not None and (matched_topic is None or matched_topic.id != topic_id):
                continue
            rows_by_topic_title[topic_title].append(
                {
                    "english_word": english_word,
                    "translation": translation,
                    "meaning_hint": _normalized_optional_str(row[3]),
                    "image_ref": _normalized_optional_str(row[5]),
                    "image_source": _normalized_optional_str(row[6]),
                    "image_prompt": _normalized_optional_str(row[7]),
                    "pixabay_search_query": _normalized_optional_str(row[8]),
                    "source_fragment": _normalized_optional_str(row[9]),
                    "is_active": _bool_from_cell(row[10]),
                }
            )

        updated_count = 0
        topic_count = 0
        content_packs: list[dict[str, object]] = []
        for topic_title, topic_rows in rows_by_topic_title.items():
            existing_topic = existing_topics_by_title.get(topic_title.casefold())
            resolved_topic_id = existing_topic.id if existing_topic is not None else _slugify(topic_title)
            existing_pack = self._store.get_content_pack(resolved_topic_id) if existing_topic is not None else None
            existing_lessons = existing_pack.get("lessons", []) if isinstance(existing_pack, dict) else []
            if not isinstance(existing_lessons, list):
                existing_lessons = []
            existing_items_by_word = _existing_items_by_english_word(
                self._store.list_vocabulary_by_topic(resolved_topic_id) if existing_topic is not None else []
            )
            content_pack = {
                "topic": {
                    "id": resolved_topic_id,
                    "title": topic_title,
                },
                "lessons": existing_lessons,
                "vocabulary_items": [
                    _build_content_pack_item(
                        topic_id=resolved_topic_id,
                        item=item,
                        existing_items_by_word=existing_items_by_word,
                        assets_dir=self._assets_dir,
                        public_asset_signing_secret=self._public_asset_signing_secret,
                        remote_image_downloader=self._remote_image_downloader,
                    )
                    for item in topic_rows
                ],
            }
            content_packs.append(content_pack)
            updated_count += len(topic_rows)
            topic_count += 1
        backup_path = _create_sqlite_backup(self._store.db_path) if content_packs else None
        self._store.upsert_content_packs_atomically(content_packs)
        return MediaCatalogImportResult(
            updated_count=updated_count,
            topic_count=topic_count,
            backup_path=backup_path,
        )


def _validate_headers(sheet, expected_headers: tuple[str, ...], *, sheet_name: str) -> None:
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if tuple(headers[: len(expected_headers)]) != expected_headers:
        raise ValueError(f"Workbook {sheet_name} headers do not match the expected format.")


def _read_workbook_topics(sheet) -> set[str]:
    topics: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        topic_title = _normalized_optional_str(row[0])
        if topic_title is None:
            continue
        folded = topic_title.casefold()
        if folded in topics:
            raise ValueError(f"Duplicate topic_title in topics sheet: {topic_title}")
        topics.add(folded)
    return topics


def _existing_items_by_english_word(items) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for item in items:
        mapping[_normalize_headword(item.english_word)] = item.id
    return mapping


def _resolve_item_id(*, topic_id: str, english_word: str, existing_items_by_word: dict[str, str]) -> str:
    normalized_word = _normalize_headword(english_word)
    existing = existing_items_by_word.get(normalized_word)
    if existing is not None:
        return existing
    base = f"{topic_id}-{_slugify(english_word)}"
    if base not in existing_items_by_word.values():
        return base
    suffix = 2
    candidate = f"{base}-{suffix}"
    while candidate in existing_items_by_word.values():
        suffix += 1
        candidate = f"{base}-{suffix}"
    return candidate


def _build_content_pack_item(
    *,
    topic_id: str,
    item: dict[str, object],
    existing_items_by_word: dict[str, str],
    assets_dir: Path,
    public_asset_signing_secret: str,
    remote_image_downloader: RemoteImageDownloader,
) -> dict[str, object]:
    item_id = _resolve_item_id(
        topic_id=topic_id,
        english_word=str(item["english_word"]),
        existing_items_by_word=existing_items_by_word,
    )
    image_ref = item.get("image_ref")
    image_source = item.get("image_source")
    if isinstance(image_ref, str) and _is_remote_url(image_ref):
        resolved_local_asset = _resolve_signed_public_image_ref(
            image_ref=image_ref,
            assets_dir=assets_dir,
            public_asset_signing_secret=public_asset_signing_secret,
        )
        if resolved_local_asset is not None:
            image_ref = resolved_local_asset
        else:
            original_image_url = image_ref
            image_ref = _download_remote_image_ref(
                url=original_image_url,
                topic_id=topic_id,
                item_id=item_id,
                assets_dir=assets_dir,
                remote_image_downloader=remote_image_downloader,
            )
            if image_source is None:
                image_source = original_image_url
    return {
        "id": item_id,
        "english_word": item["english_word"],
        "translation": item["translation"],
        **({"meaning_hint": item["meaning_hint"]} if item["meaning_hint"] is not None else {}),
        **({"image_ref": image_ref} if image_ref is not None else {}),
        **({"image_source": image_source} if image_source is not None else {}),
        **({"image_prompt": item["image_prompt"]} if item["image_prompt"] is not None else {}),
        **(
            {"pixabay_search_query": item["pixabay_search_query"]}
            if item["pixabay_search_query"] is not None
            else {}
        ),
        **({"source_fragment": item["source_fragment"]} if item["source_fragment"] is not None else {}),
        "is_active": item["is_active"],
    }


def _resolve_topic_id(store: SQLiteContentStore, item_id: str) -> str | None:
    topic_ids = store.list_topic_ids_for_item(item_id)
    return topic_ids[0] if topic_ids else None


def _is_remote_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def _download_remote_image_ref(
    *,
    url: str,
    topic_id: str,
    item_id: str,
    assets_dir: Path,
    remote_image_downloader: RemoteImageDownloader,
) -> str:
    suffix = Path(urlparse(url).path).suffix or ".jpg"
    output_path = assets_dir / topic_id / f"{item_id}{suffix}"
    remote_image_downloader.download(url=url, output_path=output_path)
    return output_path.as_posix()


def _exported_image_ref_for_item(
    *,
    image_ref: str | None,
    assets_dir: Path,
    web_app_base_url: str,
    public_asset_signing_secret: str,
) -> str | None:
    if image_ref is None:
        return None
    if _is_remote_url(image_ref):
        return image_ref
    try:
        exported_ref = build_public_asset_file_url(
            base_url=web_app_base_url,
            signing_secret=public_asset_signing_secret,
            image_ref=image_ref,
            assets_dir=assets_dir,
        )
    except (OSError, RuntimeError, ValueError):
        exported_ref = None
    return exported_ref or image_ref


def _resolve_signed_public_image_ref(
    *,
    image_ref: str,
    assets_dir: Path,
    public_asset_signing_secret: str,
) -> str | None:
    normalized_secret = public_asset_signing_secret.strip()
    if not normalized_secret:
        return None
    resolved_path = resolve_signed_public_asset_file_path(
        url=image_ref,
        signing_secret=normalized_secret,
        assets_dir=assets_dir,
    )
    if resolved_path is None:
        return None
    return resolved_path.as_posix()


def _normalized_optional_str(value: object) -> str | None:
    if value is None:
        return None
    normalized = str(value).strip()
    return normalized or None


def _normalize_headword(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return " ".join(normalized.lower().split()).strip()


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    slug = re.sub(r"[^a-z0-9]+", "-", normalized.lower().strip()).strip("-")
    return slug or "topic"


def _bool_from_cell(value: object) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value or "").strip().lower()
    if normalized in {"", "1", "true", "yes", "y"}:
        return True
    if normalized in {"0", "false", "no", "n"}:
        return False
    return True


def _bool_cell_value(value: bool) -> str:
    return "TRUE" if value else "FALSE"


def _preview_formula_for_cell(*, row_index: int, image_ref_column_index: int) -> str:
    image_ref_column = _excel_column_name(image_ref_column_index)
    return f'=IF({image_ref_column}{row_index}="","",IMAGE({image_ref_column}{row_index}))'


def _column_widths_for_headers(headers: tuple[str, ...]) -> dict[str, int]:
    widths: dict[str, int] = {}
    for index, header in enumerate(headers, start=1):
        widths[_excel_column_name(index)] = max(18, min(40, len(header) + 6))
    return widths


def _excel_column_name(index: int) -> str:
    result = ""
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _create_sqlite_backup(db_path: Path) -> Path | None:
    if not db_path.exists():
        return None
    timestamp = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    backup_dir = db_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"{db_path.stem}-pre-import-{timestamp}{db_path.suffix or '.sqlite3'}"
    source = sqlite3.connect(db_path)
    destination = sqlite3.connect(backup_path)
    try:
        source.backup(destination)
    finally:
        destination.close()
        source.close()
    return backup_path
